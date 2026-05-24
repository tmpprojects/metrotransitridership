#!/usr/bin/env python3
"""
Extract just the Metro Transit data we need (NTD ID 50027) from the FTA Monthly
Module CSVs and write it as a small JS file that index.html can load via a
plain <script> tag.

This is what lets index.html render when opened via file:// (no HTTP server
needed): browsers will load a same-folder <script src="..."> on file://, but
they won't let us fetch() a CSV.

Typical use:

    # rebuild data/metro-transit.js from whatever CSVs are already in data/
    python3 build-data.py

    # also pull the latest workbook from FTA, refresh data/*.csv, then rebuild
    python3 build-data.py --fetch

    # pull from a specific .xlsx URL (skip the FTA listing scrape)
    python3 build-data.py --fetch --url https://www.transit.dot.gov/.../...xlsx

    # only refresh CSVs from FTA, don't touch data/metro-transit.js
    python3 build-data.py --fetch --no-build

Note: --fetch needs `openpyxl` (`pip install openpyxl`).
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import gzip
import json
import re
import shutil
import subprocess
import sys
import urllib.error
import urllib.request
import zlib
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
OUTPUT = DATA_DIR / "metro-transit.js"

METRO_TRANSIT_NTD_ID = "50027"
MODES_OF_INTEREST = {"MB", "LR", "RB", "CR"}
ID_COLUMNS = {
    "NTD ID",
    "Legacy NTD ID",
    "Agency",
    "Mode/Type of Service Status",
    "Reporter Type",
    "UACE CD",
    "UZA Name",
    "Mode",
    "TOS",
    "3 Mode",
}

MONTH_RE = re.compile(r"^(\d{1,2})/(\d{4})$")

FTA_LISTING_URL = (
    "https://www.transit.dot.gov/ntd/data-product/monthly-module-adjusted-data-release"
)
# Pretend to be a browser; transit.dot.gov is fronted by a WAF that 403s the
# default urllib UA (and even a bare User-Agent header isn't always enough -- a
# full Chrome-style header set is more reliable, and we fall back to `curl` if
# urllib still gets 403'd).
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
BROWSER_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# Map normalized (alnum-lowercased) sheet name -> output CSV filename, so the
# refreshed CSVs land at the same paths build_metro_transit_js() expects.
SHEET_FILENAMES = {
    "upt": "UPT-Table.csv",
    "vrm": "VRM-Table.csv",
    "vrh": "VRH-Table.csv",
    "voms": "VOMS-Table.csv",
    "master": "Master-Table.csv",
    "calendaryearupt": "Calendar-Year-UPT.csv",
    "uptestimates": "UPT-Estimates.csv",
    "readme": "Read-Me.csv",
}


def parse_number(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = re.sub(r"[,\s]", "", value)
    if cleaned in ("", "-"):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def iso_first_of_month(header: str) -> str | None:
    m = MONTH_RE.match(header)
    if not m:
        return None
    month, year = int(m.group(1)), int(m.group(2))
    return f"{year:04d}-{month:02d}-01"


def load_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = [dict(zip(header, r)) for r in reader if r]
    return header, rows


def metro_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [r for r in rows if (r.get("NTD ID") or "").strip() == METRO_TRANSIT_NTD_ID]


def series_from_row(
    row: dict[str, str], month_cols: list[tuple[str, str]]
) -> list[dict]:
    """Build [{date, value}, ...] for one mode/row, skipping null months."""
    points = []
    for col, iso in month_cols:
        value = parse_number(row.get(col))
        if value is None:
            continue
        points.append({"date": iso, "value": value})
    return points


def build_metro_transit_js() -> None:
    upt_header, upt_rows = load_csv(DATA_DIR / "UPT-Table.csv")
    vrm_header, vrm_rows = load_csv(DATA_DIR / "VRM-Table.csv")

    upt_month_cols = [
        (col, iso_first_of_month(col))
        for col in upt_header
        if col not in ID_COLUMNS and iso_first_of_month(col) is not None
    ]
    vrm_month_cols = [
        (col, iso_first_of_month(col))
        for col in vrm_header
        if col not in ID_COLUMNS and iso_first_of_month(col) is not None
    ]

    upt_by_mode: dict[str, list[dict]] = {}
    for row in metro_rows(upt_rows):
        mode = (row.get("Mode") or "").strip()
        if mode not in MODES_OF_INTEREST:
            continue
        points = series_from_row(row, upt_month_cols)
        if not points:
            continue
        existing = upt_by_mode.get(mode, [])
        upt_by_mode[mode] = sorted(existing + points, key=lambda p: p["date"])

    bus_vrm_row = next(
        (
            r
            for r in metro_rows(vrm_rows)
            if (r.get("Mode") or "").strip() == "MB"
            and (r.get("Mode/Type of Service Status") or "").strip() == "Active"
        ),
        None,
    )
    vrm_bus = series_from_row(bus_vrm_row, vrm_month_cols) if bus_vrm_row else []

    payload = {
        "ntdId": METRO_TRANSIT_NTD_ID,
        "agency": "Metro Transit",
        "uzaName": "Minneapolis--St. Paul, MN",
        "upt": upt_by_mode,
        "vrmBus": vrm_bus,
    }

    body = json.dumps(payload, indent=2)
    OUTPUT.write_text(
        "// Auto-generated by build-data.py — do not edit by hand.\n"
        "// Re-run `python3 build-data.py` to refresh after updating CSVs in data/.\n"
        f"window.METRO_DATA = {body};\n",
        encoding="utf-8",
    )

    total_upt_points = sum(len(v) for v in upt_by_mode.values())
    print(
        f"Wrote {OUTPUT.relative_to(ROOT)} "
        f"({OUTPUT.stat().st_size:,} bytes, "
        f"{total_upt_points} UPT points across {len(upt_by_mode)} modes, "
        f"{len(vrm_bus)} Bus VRM points)."
    )


def _decode_body(data: bytes, encoding: str) -> bytes:
    encoding = (encoding or "").lower()
    if encoding == "gzip":
        return gzip.decompress(data)
    if encoding == "deflate":
        return zlib.decompress(data)
    return data


def _curl_fetch(url: str, *, accept: str) -> bytes:
    """Run `curl -sSL` and return the body. Raises if curl is missing/fails."""
    if not shutil.which("curl"):
        raise RuntimeError("curl is not on PATH; cannot fall back from urllib.")
    proc = subprocess.run(
        [
            "curl", "-sSL", "--fail", "--max-time", "180",
            "--compressed",
            "-A", USER_AGENT,
            "-H", f"Accept: {accept}",
            "-H", "Accept-Language: en-US,en;q=0.9",
            url,
        ],
        capture_output=True,
    )
    if proc.returncode != 0:
        stderr = proc.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(f"curl exited {proc.returncode}: {stderr}")
    return proc.stdout


def http_fetch(url: str, *, accept: str = "*/*") -> bytes:
    """GET url with urllib, transparently falling back to curl on 403/503."""
    headers = {**BROWSER_HEADERS, "Accept": accept}
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            return _decode_body(resp.read(), resp.headers.get("Content-Encoding", ""))
    except urllib.error.HTTPError as e:
        if e.code in (403, 503):
            print(f"urllib got HTTP {e.code} from {url}; retrying with curl...")
            return _curl_fetch(url, accept=accept)
        raise


def http_download(url: str, dest: Path) -> None:
    print(f"Downloading {url}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    data = http_fetch(
        url,
        accept=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream,*/*"
        ),
    )
    dest.write_bytes(data)
    print(f"Saved {dest.relative_to(ROOT)} ({dest.stat().st_size:,} bytes).")


def fetch_latest_xlsx_url() -> str:
    """Scrape the FTA listing page for the newest 'Complete Monthly Ridership' xlsx."""
    html = http_fetch(FTA_LISTING_URL, accept="text/html,*/*").decode(
        "utf-8", errors="replace"
    )

    # Links look like:
    #   https://www.transit.dot.gov/sites/fta.dot.gov/files/2026-05/
    #   March%202026%20Complete%20Monthly%20Ridership%20(with%20adjustments%20and%20estimates)_260501.xlsx
    pattern = re.compile(
        r'href="(?P<url>(?:https?:)?//[^"]+?Complete%20Monthly%20Ridership[^"]*?\.xlsx)"',
        re.IGNORECASE,
    )
    urls = [m.group("url") for m in pattern.finditer(html)]
    if not urls:
        raise RuntimeError(
            "Couldn't find a 'Complete Monthly Ridership' .xlsx link on "
            f"{FTA_LISTING_URL}. Open that page in a browser, copy the link to "
            "the newest .xlsx file, and re-run with `--url <that URL>`."
        )
    # First match on the page is the most recent release.
    url = urls[0]
    if url.startswith("//"):
        url = "https:" + url
    return url


def filename_for_sheet(name: str) -> str:
    key = re.sub(r"[^a-z0-9]", "", name.lower())
    if key in SHEET_FILENAMES:
        return SHEET_FILENAMES[key]
    # Fallback for any unrecognized tab: dash-separated sanitized name.
    return re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-") + ".csv"


def cell_to_str(value: object) -> str:
    """Convert an openpyxl cell value into the string form our CSVs expect."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        # Month headers like "1/2002" are stored as dates in the workbook.
        return f"{value.month}/{value.year}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def export_workbook_to_csvs(xlsx_path: Path, data_dir: Path) -> list[Path]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "--fetch needs openpyxl to read the FTA workbook. Install it with:\n"
            "    pip install openpyxl"
        ) from e

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    written: list[Path] = []
    for sheet_name in wb.sheetnames:
        out = data_dir / filename_for_sheet(sheet_name)
        ws = wb[sheet_name]
        with out.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(cell_to_str(c) for c in row)
        written.append(out)
        print(f"  Sheet {sheet_name!r:25s} -> {out.relative_to(ROOT)}")
    return written


def fetch_from_fta(explicit_url: str | None) -> None:
    try:
        url = explicit_url or fetch_latest_xlsx_url()
    except (urllib.error.URLError, RuntimeError) as e:
        raise SystemExit(
            f"Couldn't discover the latest FTA workbook URL: {e}\n"
            f"Workaround: open {FTA_LISTING_URL} in a browser, copy the link "
            "to the newest 'Complete Monthly Ridership' .xlsx, then re-run:\n"
            "    python3 build-data.py --fetch --url <that URL>"
        ) from e

    xlsx_path = DATA_DIR / "monthly-module.xlsx"
    try:
        http_download(url, xlsx_path)
    except (urllib.error.URLError, RuntimeError) as e:
        raise SystemExit(f"Download failed: {e}") from e

    print(f"Extracting sheets from {xlsx_path.relative_to(ROOT)}:")
    export_workbook_to_csvs(xlsx_path, DATA_DIR)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Refresh data/metro-transit.js from data/*.csv. With --fetch, also "
            "downloads the latest FTA Monthly Module workbook and re-exports the "
            "CSVs first."
        )
    )
    parser.add_argument(
        "--fetch",
        action="store_true",
        help="Download the latest FTA workbook and refresh data/*.csv before building.",
    )
    parser.add_argument(
        "--url",
        metavar="URL",
        help="Direct .xlsx URL to download (implies --fetch, skips FTA page scrape).",
    )
    parser.add_argument(
        "--no-build",
        action="store_true",
        help="Skip rebuilding data/metro-transit.js (useful with --fetch).",
    )
    args = parser.parse_args(argv)

    do_fetch = args.fetch or args.url is not None
    if do_fetch:
        fetch_from_fta(args.url)

    if not args.no_build:
        build_metro_transit_js()
    elif not do_fetch:
        # User passed --no-build but nothing to fetch -- nothing to do.
        print("Nothing to do: --no-build with no --fetch.", file=sys.stderr)


if __name__ == "__main__":
    main()
